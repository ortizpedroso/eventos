"""Exportação de lista de presença (PDF e Excel)."""

from __future__ import annotations

import io
from typing import Any

from fpdf import FPDF
from openpyxl import Workbook


def gerar_pdf_participantes(rows: list[dict[str, Any]], titulo: str) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, titulo[:120], ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Total: {len(rows)} registro(s)", ln=True)
    pdf.ln(4)

    cols = [
        ("Evento", 42),
        ("Participante", 38),
        ("E-mail", 48),
        ("Status", 18),
        ("Valor", 16),
        ("Check-in", 28),
    ]
    pdf.set_font("Helvetica", "B", 7)
    for label, w in cols:
        pdf.cell(w, 6, label, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for r in rows:
        checkin = r.get("checkin_em") or ""
        if checkin and "T" in str(checkin):
            checkin = str(checkin).replace("T", " ")[:16]
        vals = [
            str(r.get("evento_nome") or "")[:40],
            str(r.get("participante_nome") or "")[:35],
            str(r.get("participante_email") or "")[:45],
            str(r.get("status") or "")[:12],
            f"{float(r.get('valor') or 0):.2f}",
            str(checkin)[:24],
        ]
        for (_, w), val in zip(cols, vals, strict=True):
            pdf.cell(w, 5, val, border=1)
        pdf.ln()
        if pdf.get_y() > 270:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 7)
            for label, w in cols:
                pdf.cell(w, 6, label, border=1)
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)

    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin-1", errors="replace")
    return bytes(out)


def gerar_xlsx_participantes(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"
    headers = [
        "evento",
        "participante_nome",
        "participante_email",
        "participante_cpf",
        "participante_telefone",
        "status",
        "valor",
        "data_compra",
        "checkin_em",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r.get("evento_nome"),
                r.get("participante_nome"),
                r.get("participante_email"),
                r.get("participante_cpf"),
                r.get("participante_telefone"),
                r.get("status"),
                float(r.get("valor") or 0),
                r.get("data_compra"),
                r.get("checkin_em"),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
